import asyncio
import os
import re
from io import BytesIO

import bs4
import openai
import openpyxl
import pandas as pd
import requests
from chronological import cleaned_completion, main, read_prompt
from dotenv import load_dotenv

load_dotenv()
openai.api_key = os.getenv("OPENAI_API_KEY")

# The following are parameters for the OpenAI API
MAX_TOKENS = 800
ENGINE = "text-davinci-003"
TEMPERATURE = 0.1
TOP_P = 1
PRESENCE_PENALTY = 0

# you can name this function anything you want, the name "logic" is arbitrary
async def logic(text):
    # return await getCompletion(text)
    print('Running!')
    # you call the Chronology functions, awaiting the ones that are marked await
    prompt = read_prompt('observation-ranker').format(text)
    completion = await cleaned_completion(prompt, max_tokens=MAX_TOKENS, engine=ENGINE, temperature=TEMPERATURE, top_p=TOP_P, presence_penalty=PRESENCE_PENALTY)

    # return '{}'.format(completion)
    return completion

def getObservationRankings(raw_observations):
    return asyncio.run(logic(raw_observations))

# This fuction allows us to read an excel file from a url using openpyxl.
def load_workbook_from_url(url):
    """
    This helper function loads an excel file from a url using openpyxl. This way we don't need to download each file to our machine.
    """
    response = requests.get(url)
    wb = openpyxl.load_workbook(BytesIO(response.content))
    return wb

def getRestaurantReportCollectionLinks(year_id="#lt-229314405-2022"):
    """
    This function returns a list of URLs for weekly collections of restaurant reports. 

    Takes a year_id paramater, which you can find by inspecting the element of the year you want to scrape on the reataurant reports page. The example and default value is for 2022.
    """
    # First we get the html from the restaurant reports page.
    url = 'https://www.sanantonio.gov/Health/News/RestaurantReports'
    response = requests.get(url)
    soup = bs4.BeautifulSoup(response.text, 'html.parser')

    # Then we use the year_id to find the table of restaurant reports for that year.
    table = soup.select(f'{year_id}')

    # Then we get the links for each week of restaurant reports and add them to a list.
    linkList = []
    links = table[0].select('a')
    for link in links:
        absolute_url = 'https://www.sanantonio.gov' + link['href']
        linkList.append(absolute_url)

    return linkList

def combineWeeklyCollections():
    """
    This function combines all the weekly restaurant report collection data into one file, sorts it by total score in ascending order, and saves it as a .xlsx file.
    """
    # Start a counter to keep track of how many files we've processed.
    counter = 0

    # These are the column names we are going to rock with.
    columns = ['ESTABLISHMENT NAME', 'ESTABLISHMENT ADDRESS',
               'INSPECTION DATE',	'SECTOR',	'DISTRICT',	'TOTAL SCORE',	'LINK', 'Link']

    # Create an empty dataframe with the column names we want.
    all_data = pd.DataFrame(columns=columns)

    # Loop through each weekly collection of restaurant reports.
    for collection_link in getRestaurantReportCollectionLinks():
        
        # Add one to the counter and print it our to the terminal.
        counter += 1
        print(f"🐭 Processing file {counter}...")

        df = pd.read_excel(collection_link)

        wb = load_workbook_from_url(collection_link)
        ws = wb[wb.sheetnames[0]]

        # The following code grabs the hyperlinks from the excel file and adds them to a list.
        links = []
        for i in range(2, ws.max_row + 1):
            try:
                links.append(ws.cell(row=i, column=7).hyperlink.target)
            except:
                links.append('missing')
        
        # Now we add the list of hyperlinks to the dataframe.
        df['Link'] = links

        # Now we add the data from this week's collection to the all_data dataframe.
        all_data = pd.concat([all_data, df])
    
    # We drop the 'LINK' column because it's a duplicate of the 'Link' column.
    all_data.drop('LINK', axis=1, inplace=True)

    # We sort the data by total score in ascending order.
    all_data = all_data.sort_values(by=['TOTAL SCORE'])

    # We save the data as a .xlsx file.
    # all_data.to_excel('playground/sample.xlsx', index=False)

    return all_data


def getInspectionDetails(inspections):
    """
    This function grabs basic information about the worst inspections and isolates the observations.
    """
    # tenWorstInspections = inspections.head(10)
    # worstInspection = inspections.head(1)
    # worstInspectionLink = worstInspection['Link'].values[0]

    worstInspection = inspections.iloc[1]
    worstInspectionLink = worstInspection['Link']
    print(worstInspectionLink)

    
    response = requests.get(worstInspectionLink)
    soup = bs4.BeautifulSoup(response.text, 'html.parser')

    inspection_date = soup.find_all('td')[3].get_text().strip()[5:]
    restaurant_name = soup.find_all('td')[13].get_text().strip()[19:].title().replace('# ', '#')
    repeat_violations = soup.find_all('td')[15].find('strong').get_text().strip()[-1]
    score = soup.find_all('td')[16].get_text().strip()
    address = soup.find_all('td')[17].get_text().strip()[18:].replace('  ', ' ')

    print(restaurant_name)
    print(inspection_date)

    # Find all the table elements on the page that have a class of "padL"
    # This will give us the table elements that contain the inspection details.
    table_elements = soup.find_all('table', class_='padL')

    # Loop through each table element and add the text from each td element to a list. We will use this list to create a dataframe.
    data = []

    # These are the keywords we are looking for in the td elements.
    keywordList = ["at inspection", "observed", "encountered"]

    for table_element in table_elements:
        for td in table_element.find_all('td'):
            # If the td element text contains a keyword from the keywordList add it to the list.
            for keyword in keywordList:
                if keyword in td.get_text().lower() and "conditions observed and noted below" not in td.get_text().lower():
                    before_keyword, OG_keyword, after_keyword = td.get_text().lower().partition(keyword)
                    observation = OG_keyword + after_keyword

                    # Remove any newlines and extra spaces from the text.
                    observation = observation.replace('\n', ' ').strip()

                    # Change to sentence case.
                    observation = observation[0].upper() + observation[1:].lower()

                    # Only keep the first sentence of the observation.
                    observation = observation.split('.')[0] 

                    data.append(observation)
                else:
                    continue

    print(f'There are {len(data)} observations in this file.')

    # Join the data list with a newline character.
    data = '\n- '.join(data)

    # Add '-' to the beginning of the string.
    data = '- ' + data

    rankedObservation = getObservationRankings(data)

    inspection_writeup = f"""# {restaurant_name}

## Inspection Details

- Score: {score}
- Inspection Date: {inspection_date}
- Repeat Violations: {repeat_violations}
- Address: {address}

## Raw observations
{data}

## Ranked Observations

{rankedObservation}
"""

    # Write a markdown file with the joined data.
    with open('playground/no_2_example.md', 'w') as f:
        f = f.write(inspection_writeup)


# TODO: Remove sampleData once you're done testing. Then replace the argument in getInspectionDetails with combineWeeklyCollections().
sampleData = pd.read_excel('playground/sample.xlsx')
getInspectionDetails(sampleData)